from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QMainWindow
from PyQt5.uic import loadUi
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtCore import QTimer
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWidgets import QMainWindow, QStatusBar, QListWidget, QAction, qApp, QMenu,QMessageBox
from PyQt5.uic import loadUi
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import QSizePolicy
import openpyxl
from Database import Database
class_name = {
'1':'停车痕紧',
'2':'停车痕松',
'3':'断经',
'4':'错花',
'5':'并纬',
'6':'缩纬',
'7':'缺纬',
'8':'糙纬',
'9':'折返',
'10':'断纬',
'11':'油污',
'12':'起机',
'13':'尽机',
'14':'经条',
'15':'擦白',
'16':'擦伤',
'17':'浆斑',
'18':'空织'
}



class AddstatisForm(QMainWindow):
    def __init__(self, parent=None, ui=None):
        super(AddstatisForm, self).__init__(parent)
        loadUi("UI/Statistics.ui", self)
        self.tablemodel = QStandardItemModel(10,2)
        self.tablemodel.setHorizontalHeaderLabels(['瑕疵种类', '总数'])
        self.tablemodel_2 = QStandardItemModel(10, 8)
        self.tablemodel_2.setHorizontalHeaderLabels(['编号', '瑕疵种类', '相机', 'X坐标', 'Y坐标', 'W', 'H', '时间'])
        self.s1 = 'all'
        self.s2 = 'all'
        self.comboBox.addItem('all')
        self.comboBox.addItem('cam_01')
        self.comboBox.addItem('cam_02')
        self.comboBox.addItem('cam_03')
        self.comboBox_2.addItems(['all', '停车痕紧', '停车痕松','断经','错花','并纬','缩纬','缺纬','糙纬','折返','断纬','油污','起机','尽机','经条','擦白','擦伤','浆斑','空织'])
        self.pushButton_2.clicked.connect(self.getFile)
        self.pushButton.clicked.connect(self.sta_work_all)
        self.pushButton_3.clicked.connect(self.GetSeclectFile)
        self.pushButton_4.clicked.connect(self.select_work_all)
        self.comboBox.currentIndexChanged.connect(self.selectionchange)
        self.comboBox_2.currentIndexChanged.connect(self.selectionchange_2)
        self.cnt = -1
        self.cnt_select=-1

    def write_excel_xlsx(self,workbook,path, sheet, value):
        try:
            index = len(value)
            # workbook = openpyxl.Workbook()
            # sheet = workbook.active
            # sheet = workbook.create_sheet(str(sheet_name))
            # sheet.title = sheet_name
            for i in range(0, index):
                for j in range(0, len(value[i])):
                    sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
            workbook.save(path)
        except:
            res = QMessageBox.question(self,'警告','路径冲突，请更换路径或关闭已打开的文件')
        
    def identify_work(self, ID = 0, label = "0", camera_id = "cam_01", X = 1.0, Y = 1.0, W = 1.0, H = 1.0, flaw_time = 0):
        ID = str(ID)
        label = str(label)
        camera_id = str(camera_id)
        X = str(X)
        Y = str(Y)
        W = str(W)
        H = str(H)
        flaw_time = str(flaw_time)
        self.cnt_select += 1
        self.sm=self.tablemodel_2
        self.sm.setItem(self.cnt_select, 0, QStandardItem(ID))         
        self.sm.setItem(self.cnt_select, 1, QStandardItem(label))
        self.sm.setItem(self.cnt_select, 2, QStandardItem(camera_id))
        self.sm.setItem(self.cnt_select, 3, QStandardItem(X))
        self.sm.setItem(self.cnt_select, 4, QStandardItem(Y))
        self.sm.setItem(self.cnt_select, 5, QStandardItem(W))
        self.sm.setItem(self.cnt_select, 6, QStandardItem(H))
        self.sm.setItem(self.cnt_select, 7, QStandardItem(flaw_time))
        self.st_info.setModel(self.sm)
        #QTableView
        self.st_info.setColumnWidth(0,100)
        self.st_info.setColumnWidth(1,200)

    def downallinfo(self):
        print(book_name_xlsx)
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.title = '停车痕紧'
        if(len(book_name_xlsx)>=6 and book_name_xlsx[-5:] == '.xlsx'):
            for i in range(19)[1:]:
                li = [['编号', '类型', '相机ID', 'X轴坐标', 'Y轴坐标', '宽度', '高度', '时间戳']]
                value3 = Database.getInstance().getDataFromFlaw(flaw_type=class_name[str(i)])
                for j in value3:
                    li.append(j)
                sheet_name_xlsx = class_name[str(i)]
                if (i == 1):
                    self.write_excel_xlsx(workbook,book_name_xlsx, sheet1, li)
                else:
                    sheet = workbook.create_sheet(str(sheet_name_xlsx))
                    self.write_excel_xlsx(workbook,book_name_xlsx, sheet, li)
                li = []

    def getFile(self):
        book_name_xlsx = QFileDialog.getSaveFileName(self,
                                    "另存为",
                                    "",
                                    "Excel工作簿(*.xlsx)")[0]
        print(book_name_xlsx)
        workbook = openpyxl.Workbook()
        sheet1 = workbook.active
        sheet1.title = '停车痕紧'
        if(len(book_name_xlsx)>=6 and book_name_xlsx[-5:] == '.xlsx'):
            for i in range(19)[1:]:
                li = [['编号', '类型', '相机ID', 'X轴坐标', 'Y轴坐标', '宽度', '高度', '时间戳']]
                
                value3 = Database.getInstance().getDataFromFlaw(flaw_type=class_name[str(i)])
                for j in value3:
                    li.append(j)
                sheet_name_xlsx = class_name[str(i)]
                if (i == 1):
                    self.write_excel_xlsx(workbook,book_name_xlsx, sheet1, li)
                else:
                    sheet = workbook.create_sheet(str(sheet_name_xlsx))
                    self.write_excel_xlsx(workbook,book_name_xlsx, sheet, li)
                li = []
    
    def sta_work_all(self):                         #界面显示
        self.st_info.setModel(self.tablemodel)
        value = Database.getInstance().getallFromFlawStatistic()
        self.cnt = -1
        for i in range(len(value)):
            label = value[i][0]
            num = value[i][1]
            self.show_work(label=label, num=num)

    def show_work(self, label = "0", num=0):     
        label = str(label)
        num=str(num)
        self.cnt += 1
        self.sm=self.tablemodel
        self.sm.setItem(self.cnt, 0, QStandardItem(label))
        self.sm.setItem(self.cnt, 1, QStandardItem(num))
        self.st_info.setModel(self.sm)
        #QTableView
        self.st_info.setColumnWidth(0,100)
        self.st_info.setColumnWidth(1, 200)

    def select_work_all(self):  #界面显示
        self.st_info.setModel(self.tablemodel_2)
        str1 = self.s1
        str2 = self.s2
        li_select=[]
        if (str2 == 'all'):
            for i in range(19)[1:]:
                value3_select = Database.getInstance().getDataFromFlaw(flaw_type=class_name[str(i)])
                for j in value3_select:
                    if (j[2] == str1 or str1=='all'):
                        li_select.append(j)
        else:
            value3_select = Database.getInstance().getDataFromFlaw(flaw_type=str2)
            for j in value3_select:
                if (j[2] == str1 or str1=='all'):
                    li_select.append(j)
        # self.cnt_select = -1
        while (self.cnt_select != -1):
            self.tablemodel_2.removeRow(self.cnt_select)
            self.cnt_select -= 1
        for i in range(len(li_select)):
            ID = li_select[i][0]
            label = li_select[i][1]
            camer_id = li_select[i][2]
            X = li_select[i][3]
            Y = li_select[i][4]
            W = li_select[i][5]
            H = li_select[i][6]
            flaw_time = li_select[i][7]
            self.identify_work(ID=ID, label=label, camera_id=camer_id, X=X, Y=Y, W=W, H=H, flaw_time=flaw_time)
        
    def selectionchange(self):
        self.s1 = self.comboBox.currentText()
        print(self.s1)

    def selectionchange_2(self):
        self.s2 = self.comboBox_2.currentText()
        print(self.s2)

    def GetSeclectFile(self):
        book_name_xlsx = QFileDialog.getSaveFileName(self,
                                    "另存为",
                                    "",
                                    "Excel工作簿(*.xlsx)")[0]
        print(book_name_xlsx)
        str1 = self.s1
        str2 = self.s2
        print(str1)
        print(str2)
        if (str2 == 'all'):
            workbook = openpyxl.Workbook()
            sheet1 = workbook.active
            sheet1.title = '停车痕紧'
            if(len(book_name_xlsx)>=6 and book_name_xlsx[-5:] == '.xlsx'):
                for i in range(19)[1:]:
                    li = [['编号', '类型', '相机ID', 'X轴坐标', 'Y轴坐标', '宽度', '高度', '时间戳']]
                
                    value3 = Database.getInstance().getDataFromFlaw(flaw_type=class_name[str(i)])
                    for j in value3:
                        if (j[2] == str1 or str1=='all'):
                            li.append(j)
                    sheet_name_xlsx = class_name[str(i)]
                    if (i == 1):
                        self.write_excel_xlsx(workbook,book_name_xlsx, sheet1, li)
                    else:
                        sheet = workbook.create_sheet(str(sheet_name_xlsx))
                        self.write_excel_xlsx(workbook,book_name_xlsx, sheet, li)
                    li = []
        else:
            workbook = openpyxl.Workbook()
            sheet1 = workbook.active
            sheet1.title = str2
            if (len(book_name_xlsx) >= 6 and book_name_xlsx[-5:] == '.xlsx'):
                li = [['编号', '类型', '相机ID', 'X轴坐标', 'Y轴坐标', '宽度', '高度', '时间戳']]  
                value3 = Database.getInstance().getDataFromFlaw(flaw_type=str2)
                for j in value3:
                    if (j[2] == str1 or str1=='all'):
                        li.append(j)
                self.write_excel_xlsx(workbook,book_name_xlsx, sheet1, li)
                
        